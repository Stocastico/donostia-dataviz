"""Rent €/m² per barrio — Gobierno Vasco rental market statistics (EMA/EMAL).

Source: Observatorio Vasco de la Vivienda, "Estadística del Mercado de Alquiler"
tables (``EMAL.-Barrios-Municipios.-2016-2025_es.xlsx``), sheet **T8.3**:
average monthly rent per built m² of free-market dwellings in Donostia by
barrio, annual 2016–2024. Real registered-deposit data, not asking prices.

The EMA barrio codes (``001``–``017``) coincide with Donostia's official auzoak
codes (1–17), so we join by the integer code. Rows for which there is no
sufficient market (``-``) or that are suppressed (``.``) become missing values.
"""

from __future__ import annotations

from openpyxl import load_workbook

from ..model import BuildContext, Metric

XLSX_NAME = "emal_barrios.xlsx"
SHEET = "T8.3"
SOURCE = "Gobierno Vasco — Estadística del Mercado de Alquiler (EMA), T8.3"

# Fixed table geometry: col 1 holds the 3-digit barrio code, the nine annual
# values start at col 3 and run 2016..2024.
CODE_COL = 1
VALUE_START_COL = 3
FIRST_YEAR = 2016
N_YEARS = 9


def _to_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None  # "-" (no oferta), "." (suppressed), None → missing


def _build_from_rows(rows, code_to_id: dict[str, str]) -> list[Metric]:
    periods = [str(FIRST_YEAR + i) for i in range(N_YEARS)]
    values: dict[str, dict[str, float | None]] = {}

    for row in rows:
        code_cell = row[CODE_COL] if len(row) > CODE_COL else None
        code = str(code_cell).strip() if code_cell is not None else ""
        if not (code.isdigit() and len(code) == 3):
            continue  # only barrio rows carry a 3-digit code in this column
        barrio_id = code_to_id.get(str(int(code)))
        if not barrio_id:
            continue
        by_year = {}
        for i, period in enumerate(periods):
            col = VALUE_START_COL + i
            val = _to_float(row[col]) if len(row) > col else None
            if val is not None:
                by_year[period] = val
        if by_year:  # omit barrios with no data at all (rendered n/d)
            values[barrio_id] = by_year

    return [
        Metric(
            id="rent_eur_m2",
            label="Affitto medio (€/m²)",
            unit="€/m²",
            kind="sequential",
            theme="housing",
            source=SOURCE,
            geo_grain="barrio",
            time_grain="year",
            status="live",
            periods=periods,
            values=values,
        )
    ]


def build(ctx: BuildContext) -> list[Metric]:
    wb = load_workbook(ctx.raw_dir / XLSX_NAME, read_only=True, data_only=True)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    return _build_from_rows(rows, ctx.code_to_id)
